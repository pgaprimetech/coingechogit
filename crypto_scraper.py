"""
crypto_scraper.py
─────────────────
Scrapes CoinGecko's full coin‑market table (all pages) using Playwright,
exports the data to a styled Excel file, and emails it to you.

SETUP (one‑time)
    1.  pip install playwright openpyxl python-dotenv
    2.  playwright install chromium
    3.  cp .env.example .env          # fill in your SMTP credentials
    4.  python crypto_scraper.py

Gmail users:
    • Enable 2‑Factor Authentication on your Google account.
    • Generate an App Password at https://myaccount.google.com/apppasswords
    • Put that App Password (not your login password) into SMTP_PASS in .env
"""

from __future__ import annotations

import asyncio
import re
import smtplib
import os
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
load_dotenv()                                   # reads .env in the same folder

BASE_URL   = "https://www.coingecko.com/en/coins"   # starting page
OUTPUT_DIR = Path(__file__).resolve().parent
TIMESTAMP  = datetime.now().strftime("%Y%m%d_%H%M%S")
PAGES_DIR  = OUTPUT_DIR / "output"  # folder for individual pages and combined file

# Create the output directory
PAGES_DIR.mkdir(exist_ok=True)

# Combined Excel file will also be saved in the output folder
XLSX_PATH  = PAGES_DIR / f"coingecko_all_data_{TIMESTAMP}.xlsx"

HEADERS = [
    "Coin Name",
    "Price (USD)",
    "1h %",
    "24h %",
    "7d %",
    "24h Volume (USD)",
    "Market Cap (USD)",
    "Coin Link",
]

# ──────────────────────────────────────────────
# FILE CLEANUP
# ──────────────────────────────────────────────

def delete_old_excel_files() -> None:
    """
    Delete all old Excel files from the output directory before starting a new scrape.
    This ensures we start fresh each time and don't accumulate old files.
    """
    if not PAGES_DIR.exists():
        return
    
    # Find all Excel files in the output directory
    excel_files = list(PAGES_DIR.glob("*.xlsx"))
    
    if not excel_files:
        print("  🗑️  No old Excel files to delete.\n")
        return
    
    deleted_count = 0
    failed_count = 0
    
    print(f"  🗑️  Found {len(excel_files)} old Excel file(s) to delete...")
    
    for file_path in excel_files:
        try:
            # Skip temporary Excel files (start with ~$)
            if file_path.name.startswith("~$"):
                continue
            
            file_path.unlink()  # Delete the file
            print(f"     ✓ Deleted: {file_path.name}")
            deleted_count += 1
        except PermissionError:
            print(f"     ✗ Cannot delete (file is open): {file_path.name}")
            failed_count += 1
        except Exception as e:
            print(f"     ✗ Error deleting {file_path.name}: {e}")
            failed_count += 1
    
    if deleted_count > 0:
        print(f"  ✅  Deleted {deleted_count} old file(s).\n")
    if failed_count > 0:
        print(f"  ⚠️  Failed to delete {failed_count} file(s) (may be open in Excel).\n")


# ──────────────────────────────────────────────
# SCRAPER
# ──────────────────────────────────────────────

async def scrape_page(page) -> list[list[str]]:
    """
    Parse every visible coin‑row on the current CoinGecko page.
    Handles both the classic <table> layout and the newer div‑based layout.
    """
    rows_data: list[list[str]] = []

    def clean_coin_name(name: str) -> str:
        """Remove rank numbers, 'Buy' buttons, and extra whitespace from coin names."""
        # Split by newlines and filter out unwanted parts
        parts = [p.strip() for p in name.split("\n") if p.strip()]
        
        # Remove numeric ranks, 'Buy' text, and other noise
        cleaned_parts = []
        for part in parts:
            # Skip if it's just a number (rank)
            if part.replace(",", "").replace(".", "").isdigit():
                continue
            # Skip 'Buy' button text
            if part.lower() == "buy":
                continue
            # Skip very short parts that are likely icons/buttons
            if len(part) <= 2 and not part.isalpha():
                continue
            cleaned_parts.append(part)
        
        # Join remaining parts with space
        return " ".join(cleaned_parts).strip()

    # First, let's wait a bit longer for the page to fully load
    await page.wait_for_timeout(3000)
    
    print("    🔍 Analyzing page structure...")
    
    # Debug: Check what's on the page
    page_content = await page.evaluate("""() => {
        return {
            hasTables: document.querySelectorAll('table').length,
            hasTableRows: document.querySelectorAll('table tbody tr').length,
            hasDivRows: document.querySelectorAll('[data-testid="table-row"]').length,
            bodyText: document.body.innerText.substring(0, 500)
        };
    }""")
    
    print(f"    📊 Found {page_content['hasTables']} tables, {page_content['hasTableRows']} table rows, {page_content['hasDivRows']} div rows")

    # ── Strategy A: Use JavaScript to properly extract structured data from tables ──
    js_data = await page.evaluate("""() => {
        const results = [];
        
        // Find all table rows
        const rows = document.querySelectorAll("table tbody tr");
        
        console.log('Found rows:', rows.length);
        
        rows.forEach((row, index) => {
            const cells = row.querySelectorAll("td");
            console.log(`Row ${index}: ${cells.length} cells`);
            
            if (cells.length < 7) return;
            
            const data = {};
            
            // Get all cell text values first
            const cellTexts = Array.from(cells).map(cell => cell.innerText.trim());
            
            // Try to find the coin link to determine structure
            const coinLink = row.querySelector('a[href*="/coins/"]');
            if (coinLink) {
                data.coinUrl = coinLink.href;
                data.name = coinLink.innerText.trim();
            } else {
                // Fallback: use second cell for name (first is usually rank)
                data.name = cellTexts[1] || '';
                data.coinUrl = '';
            }
            
            // Try to identify which cells contain what based on content patterns
            let dataIndex = 0;
            for (let i = 0; i < cellTexts.length; i++) {
                const text = cellTexts[i];
                
                // Skip rank numbers
                if (i === 0 || (text && text.match(/^\d+$/))) continue;
                
                // Skip coin name (already got it)
                if (text === data.name) continue;
                
                // Price (contains $ or starts with number with decimal)
                if (!data.price && (text.includes('$') || text.match(/^\d+[\.,]\d/))) {
                    data.price = text;
                    continue;
                }
                
                // Percentage changes (contain %)
                if (text.includes('%')) {
                    if (!data.change1h) data.change1h = text;
                    else if (!data.change24h) data.change24h = text;
                    else if (!data.change7d) data.change7d = text;
                    continue;
                }
                
                // Volume and market cap (large numbers with $ or B/M/K)
                if (text.match(/[\$\d]/) && text.match(/[BMK]|billion|million/i)) {
                    if (!data.volume) data.volume = text;
                    else if (!data.marketCap) data.marketCap = text;
                    continue;
                }
                
                // If it's a number with commas or large number
                if (text.match(/^\$?[\d,]+(\.\d+)?[BMK]?$/i)) {
                    if (!data.volume) data.volume = text;
                    else if (!data.marketCap) data.marketCap = text;
                }
            }
            
            // Set defaults for missing data
            data.change1h = data.change1h || '';
            data.change24h = data.change24h || '';
            data.change7d = data.change7d || '';
            data.volume = data.volume || '';
            data.marketCap = data.marketCap || '';
            data.graphLink = data.coinUrl || '';
            
            if (data.name) {
                results.push(data);
            }
        });
        
        return results;
    }""")

    print(f"    ✅ JavaScript extraction found {len(js_data)} coins")

    for item in js_data:
        coin_name = clean_coin_name(item.get('name', ''))
        if not coin_name:
            continue
            
        rows_data.append([
            coin_name,
            item.get('price', ''),
            item.get('change1h', ''),
            item.get('change24h', ''),
            item.get('change7d', ''),
            item.get('volume', ''),
            item.get('marketCap', ''),
            item.get('graphLink', ''),
        ])

    # ── Strategy B: Fallback - get all text from rows and parse manually ──
    if not rows_data:
        print("    ⚠️  Trying fallback method...")
        
        fallback_data = await page.evaluate("""() => {
            const results = [];
            const rows = document.querySelectorAll("table tbody tr");
            
            rows.forEach(row => {
                const allText = row.innerText;
                const link = row.querySelector('a[href*="/coins/"]');
                
                results.push({
                    text: allText,
                    url: link ? link.href : ''
                });
            });
            
            return results;
        }""")
        
        print(f"    📝 Fallback found {len(fallback_data)} rows of raw text")
        
        for item in fallback_data:
            text = item.get('text', '')
            url = item.get('url', '')
            
            # Split by tabs and newlines
            parts = [p.strip() for p in text.replace('\t', '\n').split('\n') if p.strip()]
            
            if len(parts) < 7:
                continue
            
            # Filter out rank and Buy
            filtered = []
            for part in parts:
                if part.replace(',', '').replace('.', '').isdigit() and len(part) < 4:
                    continue
                if part.lower() == 'buy':
                    continue
                filtered.append(part)
            
            if len(filtered) < 7:
                continue
            
            coin_name = clean_coin_name(filtered[0])
            if not coin_name:
                continue
            
            rows_data.append([
                coin_name,
                filtered[1] if len(filtered) > 1 else '',
                filtered[2] if len(filtered) > 2 else '',
                filtered[3] if len(filtered) > 3 else '',
                filtered[4] if len(filtered) > 4 else '',
                filtered[5] if len(filtered) > 5 else '',
                filtered[6] if len(filtered) > 6 else '',
                url,
            ])

    print(f"    📦 Total rows extracted: {len(rows_data)}")
    return rows_data


async def click_next_page(page) -> bool:
    """
    Attempts to click the "Next" pagination button using multiple strategies.
    Returns True if successfully clicked, False if no next button found.
    """
    # Strategy 1: Look for the right arrow icon with cursor pointer
    try:
        next_button = await page.query_selector('i.fa-angle-right.tw-cursor-pointer')
        if next_button:
            # Check if it's disabled (has opacity class)
            is_disabled = await page.evaluate("""(el) => {
                const classList = Array.from(el.classList);
                return classList.some(c => c.startsWith('tw-opacity-'));
            }""", next_button)
            
            if not is_disabled:
                # Click the parent element (usually the button/link container)
                parent = await next_button.evaluate_handle('el => el.parentElement')
                await parent.click()
                print("    ✓ Clicked next button (Strategy 1: icon)")
                await page.wait_for_timeout(2000)  # Wait for page to load
                return True
    except Exception as e:
        print(f"    Strategy 1 failed: {e}")

    # Strategy 2: Look for clickable pagination elements
    try:
        # Try to find a clickable pagination arrow/button
        selectors = [
            'a[rel="next"]',
            'button[aria-label*="Next"]',
            'button[aria-label*="next"]',
            '[class*="pagination"] a:last-child',
            '[class*="pagination"] button:last-child',
        ]
        
        for selector in selectors:
            element = await page.query_selector(selector)
            if element:
                is_disabled = await element.evaluate(
                    '(el) => el.disabled || el.classList.contains("disabled") || el.getAttribute("aria-disabled") === "true"'
                )
                if not is_disabled:
                    await element.click()
                    print(f"    ✓ Clicked next button (Strategy 2: {selector})")
                    await page.wait_for_timeout(2000)
                    return True
    except Exception as e:
        print(f"    Strategy 2 failed: {e}")

    # Strategy 3: JavaScript click on the icon's parent
    try:
        clicked = await page.evaluate("""() => {
            const icons = document.querySelectorAll('i.fa-angle-right.tw-cursor-pointer');
            for (const icon of icons) {
                const classList = Array.from(icon.classList);
                const isDisabled = classList.some(c => c.startsWith('tw-opacity-'));
                if (!isDisabled && icon.parentElement) {
                    icon.parentElement.click();
                    return true;
                }
            }
            return false;
        }""")
        
        if clicked:
            print("    ✓ Clicked next button (Strategy 3: JS click)")
            await page.wait_for_timeout(2000)
            return True
    except Exception as e:
        print(f"    Strategy 3 failed: {e}")

    # Strategy 4: Find any element containing the right arrow and click it
    try:
        clicked = await page.evaluate("""() => {
            // Look for the FontAwesome right arrow icon
            const allElements = Array.from(document.querySelectorAll('*'));
            for (const el of allElements) {
                if (el.classList.contains('fa-angle-right') || 
                    el.classList.contains('fa-chevron-right') ||
                    el.classList.contains('fa-arrow-right')) {
                    
                    // Check if clickable (has cursor pointer)
                    const style = window.getComputedStyle(el);
                    if (style.cursor === 'pointer') {
                        // Click the element or its parent
                        const clickable = el.closest('a, button') || el.parentElement;
                        if (clickable) {
                            clickable.click();
                            return true;
                        }
                    }
                }
            }
            return false;
        }""")
        
        if clicked:
            print("    ✓ Clicked next button (Strategy 4: comprehensive search)")
            await page.wait_for_timeout(2000)
            return True
    except Exception as e:
        print(f"    Strategy 4 failed: {e}")

    return False


async def scrape_all() -> list[list[str]]:
    """
    Launch Playwright, iterate through every paginated page on CoinGecko,
    and collect all coin rows. Saves all data progressively to a single Excel file
    that gets updated after each page is scraped.
    """
    all_rows: list[list[str]] = []
    page_num = 1

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1280, "height": 800},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()

        # Block ad / tracker domains so background connections don't stall us
        BLOCKED = [
            "doubleclick.net", "googlesyndication.com", "adsbygoogle",
            "googletagmanager.com", "facebook.net", "hotjar.com",
            "interstitial", "ads.coingecko", "adtarget", "pubmatic",
        ]
        await page.route(
            "**/*",
            lambda route: (
                route.abort()
                if any(b in route.request.url for b in BLOCKED)
                else route.continue_()
            ),
        )

        url = BASE_URL                      # first page (no ?page= param)

        while True:
            print(f"  ► Scraping page {page_num}  …  {url}")
            
            # Use 'domcontentloaded' instead of 'networkidle' to avoid timeout issues
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=60_000)
                # Give a bit more time for dynamic content to load
                await page.wait_for_timeout(2000)
            except Exception as e:
                print(f"    ⚠️  Navigation error: {e}")
                print("    🔄 Retrying with 'load' wait state...")
                try:
                    await page.goto(url, wait_until="load", timeout=60_000)
                    await page.wait_for_timeout(2000)
                except Exception as e2:
                    print(f"    ✗  Navigation failed completely: {e2}")
                    break

            # Wait longer for the coin‑row table to render
            print("    ⏳ Waiting for table to load...")
            try:
                await page.wait_for_selector(
                    'table tbody tr, [data-testid="table-row"]',
                    timeout=20_000
                )
                print("    ✓ Table found!")
            except Exception as e:
                print(f"    ⚠️  Table selector timeout: {e}")
                print("    ⏳ Waiting additional 5 seconds...")
                await page.wait_for_timeout(5_000)

            rows = await scrape_page(page)
            if not rows:
                print(f"    ✗ No rows found on page {page_num}")
                
                # Take a screenshot for debugging
                screenshot_path = OUTPUT_DIR / f"debug_page_{page_num}.png"
                await page.screenshot(path=str(screenshot_path))
                print(f"    📸 Screenshot saved to {screenshot_path.name} for debugging")
                
                # Try one more time with a longer wait
                if page_num == 1:
                    print("    🔄 Retrying page 1 with longer wait...")
                    await page.wait_for_timeout(10_000)
                    rows = await scrape_page(page)
                    
                if not rows:
                    print("    ✗ Still no rows found – stopping.")
                    break

            all_rows.extend(rows)
            print(f"    ✓ {len(rows)} coins collected  (total so far: {len(all_rows)})")

            # ── Update the single Excel file with all data collected so far ──
            print(f"    💾 Updating Excel file with {len(all_rows)} total coins...")
            build_excel(all_rows)

            # ── Try to click the next page button ──
            print("    → Attempting to navigate to next page...")
            success = await click_next_page(page)
            
            if not success:
                print("    ✗ Could not find or click next button – all pages scraped.")
                break
            
            # Wait for the new page content to load
            try:
                await page.wait_for_selector(
                    'table tbody tr, [data-testid="table-row"]',
                    timeout=10_000
                )
            except Exception:
                await page.wait_for_timeout(3_000)
            
            page_num += 1
            url = page.url  # Update URL to reflect current page

        await browser.close()

    return all_rows


# ──────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────

# Colour palette
DARK_BG      = PatternFill("solid", fgColor="1A1A2E")   # header bg
GOLD_FONT    = Font(name="Arial", bold=True, size=11, color="FFD700")
WHITE_FONT   = Font(name="Arial", size=10, color="FFFFFF")
BLACK_FONT   = Font(name="Arial", size=10, color="1A1A2E")
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F7FC")
THIN_BORDER  = Border(
    left  =Side(style="thin", color="D0D8E0"),
    right =Side(style="thin", color="D0D8E0"),
    top   =Side(style="thin", color="D0D8E0"),
    bottom=Side(style="thin", color="D0D8E0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def build_excel(rows: list[list[str]]) -> Path:
    """
    Create a professionally styled .xlsx workbook from the scraped rows.
    Overwrites the file each time to progressively add more data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CoinGecko Data"

    # ── Title row ──
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value   = "CoinGecko – Cryptocurrency Market Data"
    title_cell.font    = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    title_cell.fill    = DARK_BG
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Subtitle / timestamp ──
    ws.merge_cells("A2:H2")
    sub_cell = ws["A2"]
    sub_cell.value   = f"Scraped on  {datetime.now().strftime('%d %b %Y, %H:%M')}  •  {len(rows)} coins"
    sub_cell.font    = Font(name="Arial", italic=True, size=10, color="6B7280")
    sub_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # ── Header row (row 3) ──
    header_row = 3
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value      = header
        cell.font       = GOLD_FONT
        cell.fill       = DARK_BG
        cell.alignment  = CENTER
        cell.border     = THIN_BORDER
    ws.row_dimensions[header_row].height = 22

    # ── Data rows (start at row 4) ──
    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        is_alt = (row_idx % 2 == 0)           # alternating row shading
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value     = value
            cell.font      = BLACK_FONT
            cell.border    = THIN_BORDER

            # Coin name – left‑aligned; everything else centred
            cell.alignment = LEFT if col_idx == 1 else CENTER

            # Alternate row fill
            if is_alt:
                cell.fill = ALT_ROW_FILL

        ws.row_dimensions[row_idx].height = 20

    # ── Column widths (hand‑tuned) ──
    col_widths = {
        1: 34,   # Coin Name
        2: 16,   # Price
        3: 10,   # 1h
        4: 10,   # 24h
        5: 10,   # 7d
        6: 20,   # 24h Volume
        7: 22,   # Market Cap
        8: 50,   # Coin Link
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze header row so it sticks when scrolling ──
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Sheet protection (optional: read‑only feel) – commented out so user can edit ──
    # ws.protection.sheet = True

    wb.save(str(XLSX_PATH))
    # Don't print on every save to keep output clean - scraper prints the updates
    return XLSX_PATH


# ──────────────────────────────────────────────
# EMAIL
# ──────────────────────────────────────────────

def send_email(xlsx_path: Path) -> None:
    """
    Attach the Excel file to an email and send it via SMTP (TLS).
    Credentials are read from .env through dotenv.
    """
    smtp_host      = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port      = int(os.getenv("SMTP_PORT", "587"))
    smtp_user      = os.getenv("SMTP_USER", "")
    smtp_pass      = os.getenv("SMTP_PASS", "")
    recipient      = os.getenv("RECIPIENT_EMAIL", "")

    if not all([smtp_user, smtp_pass, recipient]):
        print("\n  ⚠  EMAIL SKIPPED – fill SMTP_USER, SMTP_PASS, "
              "and RECIPIENT_EMAIL in .env to enable emailing.")
        return

    # ── build MIME message ──
    msg            = MIMEMultipart()
    msg["From"]    = smtp_user
    msg["To"]      = recipient
    msg["Subject"] = f"CoinGecko Data Export – {datetime.now().strftime('%d %b %Y')}"

    body_text = (
        "Hi,\n\n"
        "Please find attached the latest CoinGecko cryptocurrency market data.\n"
        f"Total coins scraped: see the subtitle row inside the Excel file.\n\n"
        "Best,\nCrypto Scraper Bot\n"
    )
    msg.attach(MIMEText(body_text, "plain"))

    # ── attach Excel ──
    with open(xlsx_path, "rb") as f:
        attachment = MIMEApplication(f.read(), Name=xlsx_path.name)
    attachment["Content-Disposition"] = f'attachment; filename="{xlsx_path.name}"'
    msg.attach(attachment)

    # ── send ──
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, recipient, msg.as_string())
        print(f"  ✉️  Email sent to  {recipient}")
    except Exception as exc:
        print(f"  ✗  Email failed: {exc}")
        print("     → Double‑check SMTP credentials in .env")


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

async def main():
    print("\n╔══════════════════════════════════════════╗")
    print("║   CoinGecko Scraper  –  Playwright       ║")
    print("╚══════════════════════════════════════════╝\n")

    print(f"  📁  All files will be saved to: output/\n")
    
    # Delete old Excel files before starting
    delete_old_excel_files()
    
    print(f"  📊  File will be updated progressively: {XLSX_PATH.name}\n")

    print("  📥  Phase 1 – Scraping (file updates after each page) …")
    rows = await scrape_all()

    if not rows:
        print("\n  ✗  No data was collected. Exiting.")
        return

    print(f"\n  ✅  Scraping complete! Total: {len(rows)} coins")
    print(f"  📄  Final file: {XLSX_PATH.name}")

    print(f"\n  📧  Phase 2 – Sending email …")
    send_email(XLSX_PATH)

    print("\n  ✅  Done!")
    print(f"  📁  Output folder: {PAGES_DIR}")
    print(f"  📄  File: {XLSX_PATH.name}\n")


if __name__ == "__main__":
    asyncio.run(main())